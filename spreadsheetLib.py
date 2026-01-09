from config import *
import openpyxl as xlsx

xlsx.load_workbook("template.xlsx").save("result.xlsx")
spreadsheet = xlsx.load_workbook("result.xlsx")
data_sheet = spreadsheet[config["spreadsheet-names"]["data"]]


def getNumCols(sheet=data_sheet) -> int:
    n_c = 0
    row_dat = next(sheet.iter_rows(1, 1))
    while row_dat[n_c].value: n_c += 1
    if n_c > 0: n_c -= 1

    return n_c


def getColNames(sheet=data_sheet) -> list[str]:
    return [n.value for n in next(sheet.iter_rows(1, 1))[:getNumCols(sheet)]]


num_cols = getNumCols()
col_names = {c_n.replace(" ", "_"): i for i, c_n in enumerate(getColNames())}
data_sheet.delete_cols(num_cols + 2, 2 ** 31)


class Row(list):
    def __init__(self, data: list = None):
        if data is None:
            data = [None for _ in range(num_cols)]
        super().__init__(data)

    def update(self, **data):
        for key, val in data.items():
            self[col_names[key]] = val


def insertRow(data: Row, sheet=data_sheet):
    if len(data) != num_cols:
        raise IndexError("row must have the same number of columns as template.xlsx")
    sheet.append(data)


def save():
    spreadsheet.save("result.xlsx")


if __name__ == "__main__":
    print(col_names)
